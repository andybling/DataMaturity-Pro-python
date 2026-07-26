"""Génération des livrables PDF et Excel."""

from datetime import datetime, timezone

from app.data.grid import ALL_CRITERIA
from app.models import Assessment
from app.services.analysis import build_analysis
from app.services.reports import build_grid_xlsx, build_report_pdf, report_filename
from app.services.scoring import compute_score


def _fixture():
    assessment = Assessment(
        company_name="Rapport Test SA",
        sector="Industrie & Manufacturing",
        country="Côte d'Ivoire",
        company_size="200-999",
        contact_name="Test",
        contact_email="test@example.com",
        created_at=datetime.now(timezone.utc),
    )
    assessment.answers = {c.code: (i * 5) % 4 for i, c in enumerate(ALL_CRITERIA)}
    analysis = build_analysis(
        compute_score(assessment.answers),
        company_size_code="200-999",
        company_name=assessment.company_name,
        sector=assessment.sector,
    )
    return assessment, analysis


def test_pdf_premium_plus_riche_que_standard():
    assessment, analysis = _fixture()
    standard = build_report_pdf(assessment, analysis, plan_code="standard")
    premium = build_report_pdf(assessment, analysis, plan_code="premium")
    assert standard.startswith(b"%PDF")
    assert premium.startswith(b"%PDF")
    assert len(premium) > len(standard)


def test_export_excel_et_plan_daction():
    import io

    from openpyxl import load_workbook

    assessment, analysis = _fixture()
    payant = load_workbook(io.BytesIO(build_grid_xlsx(assessment, analysis)))
    gratuit = load_workbook(
        io.BytesIO(build_grid_xlsx(assessment, analysis, include_action_plan=False))
    )
    assert "Plan d'action" in payant.sheetnames
    assert "Plan d'action" not in gratuit.sheetnames
    assert payant["Grille détaillée"].max_row == 46  # en-tête + 45 critères


def test_nom_de_fichier_assaini():
    assessment, _ = _fixture()
    assessment.company_name = "Société / Test & Co"
    name = report_filename(assessment, "pdf", "premium")
    assert name.endswith(".pdf")
    assert "/" not in name and "&" not in name


def test_radar_svg_valide():
    from app.services.charts import radar_svg

    svg = radar_svg(["A", "B", "C", "D", "E", "F", "G"], [10, 20, 30, 40, 50, 60, 70])
    assert svg.startswith("<svg") and svg.endswith("</svg>")
    assert svg.count("<polygon") >= 5
