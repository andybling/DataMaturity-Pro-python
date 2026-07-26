# -*- coding: utf-8 -*-
import openpyxl, json, re, unicodedata

"""Extraction de la grille Limpida depuis le fichier Excel source.

Usage :
    python scripts/extract_grid_from_xlsx.py "Grille de maturité Data_Limpida_2024.xlsx"

Produit grid.json, consommé ensuite par scripts/generate_grid.py pour régénérer
app/data/grid.py. À relancer uniquement si Limpida publie une nouvelle version
de la grille.
"""

import sys

XL = sys.argv[1] if len(sys.argv) > 1 else 'Grille de maturité Data_Limpida_2024.xlsx' 
wb = openpyxl.load_workbook(XL, data_only=True)
ws_desc = wb['Évaluation des critères']
ws_ev = wb['Votre évaluation']

def clean(v):
    if v is None: return None
    s = str(v).replace('\n', ' ')
    s = re.sub(r'\s+', ' ', s).strip()
    return s or None

# --- 1. descriptions per criterion
dims = []          # [{name, criteria:[{name, levels:[4]}]}]
cur = None
for row in ws_desc.iter_rows(min_row=6, max_row=59):
    theme = clean(row[1].value)   # B
    crit  = clean(row[2].value)   # C
    lv = [clean(row[i].value) for i in range(3, 7)]  # D..G
    if theme:
        cur = {'name': theme, 'criteria': []}
        dims.append(cur)
    if crit and cur is not None:
        cur['criteria'].append({'name': crit, 'levels': lv})

# --- 2. weights
IMP = {'3': 3, '2': 2, '1': 1}
def parse_w(v):
    s = clean(v)
    if not s: return None
    m = re.match(r'^(\d)', s)
    return int(m.group(1)) if m else None

ev_dims = []
cur = None
for row in ws_ev.iter_rows(min_row=6, max_row=62):
    theme = clean(row[1].value)     # B theme
    tw    = parse_w(row[2].value)   # C theme weight
    crit  = clean(row[3].value)     # D criterion
    cw    = parse_w(row[4].value)   # E criterion weight
    mx    = row[5].value            # F max score
    if theme and theme.lower().startswith('total'): break
    if theme and tw:
        cur = {'name': theme, 'weight': tw, 'criteria': []}
        ev_dims.append(cur)
    if crit and cw and cur is not None and not str(crit).lower().startswith('sous total'):
        cur['criteria'].append({'name': crit, 'weight': cw, 'max': int(mx) if isinstance(mx,(int,float)) else None})

# --- merge (order is identical between sheets)
assert len(dims) == len(ev_dims), (len(dims), len(ev_dims))
out = []
for d, e in zip(dims, ev_dims):
    assert len(d['criteria']) == len(e['criteria']), (d['name'], len(d['criteria']), len(e['criteria']))
    crits = []
    for c, ec in zip(d['criteria'], e['criteria']):

        crits.append({
            'name': c['name'], 'levels': c['levels'],
            'weight': ec['weight'], 'max': ec['max'],
        })
    out.append({'name': d['name'], 'weight': e['weight'], 'criteria': crits})

total = 0
for d in out:
    sub = 0
    for c in d['criteria']:
        calc = 3 * c['weight'] * d['weight']
        assert calc == c['max'], (d['name'], c['name'], calc, c['max'])
        sub += calc
    total += sub
    print(f"{d['name']}: poids={d['weight']} crit={len(d['criteria'])} sous-total={sub}")
print('TOTAL', total, 'criteres', sum(len(d['criteria']) for d in out))
json.dump(out, open('grid.json','w'), ensure_ascii=False, indent=1)
