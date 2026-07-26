"""Exports de pilotage : CSV et Excel de la base de prospects."""

from __future__ import annotations

import csv
import io
from typing import Iterable, List

from app.models import ORDER_PAID, Assessment

COLUMNS = [
    "public_id", "date", "organisation", "secteur", "pays", "effectif",
    "contact", "fonction", "email", "telephone", "canal",
    "statut", "etape_commerciale", "score", "maximum", "pourcentage", "niveau",
    "gouvernance", "qualite", "securite", "integration", "analyse", "culture", "infrastructure",
    "commandes_payees", "montant_paye_fcfa", "offre",
]

DIM_ORDER = ["governance", "quality", "security", "integration", "analytics", "culture", "infrastructure"]


def _row(assessment: Assessment) -> List:
    scores = assessment.dimension_scores
    paid = [o for o in assessment.orders if o.status == ORDER_PAID]
    return [
        assessment.public_id,
        (assessment.completed_at or assessment.created_at).strftime("%Y-%m-%d %H:%M"),
        assessment.company_name,
        assessment.sector,
        assessment.country,
        assessment.company_size,
        assessment.contact_name,
        assessment.contact_role,
        assessment.contact_email,
        assessment.contact_phone,
        assessment.acquisition_channel,
        assessment.status,
        assessment.lead_stage,
        assessment.total_score,
        assessment.max_score,
        round(assessment.percentage, 1),
        assessment.level_code,
        *[scores.get(code, {}).get("percentage", "") for code in DIM_ORDER],
        len(paid),
        sum(o.amount_xof for o in paid),
        assessment.paid_plan_code or "",
    ]


def to_csv(rows: Iterable[Assessment]) -> bytes:
    buffer = io.StringIO()
    writer = csv.writer(buffer, delimiter=";", quoting=csv.QUOTE_MINIMAL)
    writer.writerow(COLUMNS)
    for assessment in rows:
        writer.writerow(_row(assessment))
    # BOM UTF-8 : Excel en environnement francophone ouvre le fichier correctement.
    return b"\xef\xbb\xbf" + buffer.getvalue().encode("utf-8")


def to_xlsx(rows: Iterable[Assessment]) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Prospects"
    header_font = Font(bold=True, color="FFFFFF", size=10)
    header_fill = PatternFill("solid", fgColor="0F172A")
    for col, name in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col, value=name)
        cell.font, cell.fill = header_font, header_fill
        cell.alignment = Alignment(wrap_text=True, vertical="center")
    for index, assessment in enumerate(rows, start=2):
        for col, value in enumerate(_row(assessment), start=1):
            ws.cell(row=index, column=col, value=value)
    widths = [14, 17, 30, 24, 16, 14, 22, 20, 30, 16, 18, 12, 16, 8, 9, 12, 12,
              12, 10, 10, 12, 9, 9, 14, 10, 18, 12]
    for col, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}{max(1, ws.max_row)}"

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
