"""Génération des livrables : PDF professionnel et export Excel de la grille."""

from __future__ import annotations

import io
from datetime import datetime, timezone
from typing import List, Optional

from reportlab.lib import colors
from reportlab.lib.enums import TA_JUSTIFY, TA_RIGHT
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.pdfgen import canvas as pdfcanvas
from reportlab.platypus import (
    BaseDocTemplate,
    Frame,
    KeepTogether,
    PageBreak,
    PageTemplate,
    Paragraph,
    Spacer,
    Table,
    TableStyle,
)

from app.config import settings
from app.data.grid import DIMENSIONS
from app.models import Assessment
from app.services.analysis import Analysis
from app.services.charts import axis_points, radar_points
from app.services.pricing import format_xof

INK = colors.HexColor("#0F172A")
MUTED = colors.HexColor("#64748B")
LINE = colors.HexColor("#E2E8F0")
BRAND = colors.HexColor("#4F46E5")
SOFT = colors.HexColor("#F1F5F9")


# ---------------------------------------------------------------------------
#  Styles
# ---------------------------------------------------------------------------


def _styles() -> dict:
    base = getSampleStyleSheet()
    return {
        "title": ParagraphStyle("t", parent=base["Title"], fontName="Helvetica-Bold",
                                fontSize=26, leading=30, textColor=INK, spaceAfter=6),
        "subtitle": ParagraphStyle("st", parent=base["Normal"], fontSize=12, leading=17,
                                   textColor=MUTED, spaceAfter=18),
        "h1": ParagraphStyle("h1", parent=base["Heading1"], fontName="Helvetica-Bold",
                             fontSize=16, leading=20, textColor=INK, spaceBefore=16, spaceAfter=8),
        "h2": ParagraphStyle("h2", parent=base["Heading2"], fontName="Helvetica-Bold",
                             fontSize=12, leading=16, textColor=BRAND, spaceBefore=12, spaceAfter=5),
        "body": ParagraphStyle("b", parent=base["Normal"], fontSize=9.6, leading=14.5,
                               textColor=INK, alignment=TA_JUSTIFY, spaceAfter=7),
        "small": ParagraphStyle("s", parent=base["Normal"], fontSize=8.2, leading=11.5,
                                textColor=MUTED, spaceAfter=4),
        "cell": ParagraphStyle("c", parent=base["Normal"], fontSize=8.4, leading=11.5, textColor=INK),
        "cellb": ParagraphStyle("cb", parent=base["Normal"], fontSize=8.4, leading=11.5,
                                textColor=INK, fontName="Helvetica-Bold"),
        # En-tête de tableau : fond sombre, donc texte blanc. La propriété
        # TEXTCOLOR de TableStyle n'agit pas sur le texte d'un Paragraph.
        "cellh": ParagraphStyle("ch", parent=base["Normal"], fontSize=8.2, leading=11,
                                textColor=colors.white, fontName="Helvetica-Bold"),
        "cellhr": ParagraphStyle("chr", parent=base["Normal"], fontSize=8.2, leading=11,
                                 textColor=colors.white, fontName="Helvetica-Bold",
                                 alignment=TA_RIGHT),
        "cellr": ParagraphStyle("cr", parent=base["Normal"], fontSize=8.4, leading=11.5,
                                textColor=INK, alignment=TA_RIGHT),
        "kpi": ParagraphStyle("k", parent=base["Normal"], fontSize=22, leading=25,
                              fontName="Helvetica-Bold", textColor=BRAND),
    }


class _Doc(BaseDocTemplate):
    """Gabarit avec en-tête, pied de page et pagination."""

    def __init__(self, buffer, *, title: str, company: str, **kw):
        super().__init__(buffer, pagesize=A4, title=title, author=settings.brand_owner,
                         leftMargin=18 * mm, rightMargin=18 * mm,
                         topMargin=24 * mm, bottomMargin=20 * mm, **kw)
        self.company = company
        self.doc_title = title
        frame = Frame(self.leftMargin, self.bottomMargin, self.width, self.height, id="main")
        self.addPageTemplates([PageTemplate(id="std", frames=[frame], onPage=self._decorate)])

    def _decorate(self, cv: pdfcanvas.Canvas, doc) -> None:
        cv.saveState()
        page = cv.getPageNumber()
        if page > 1:
            cv.setFont("Helvetica-Bold", 8)
            cv.setFillColor(BRAND)
            cv.drawString(doc.leftMargin, A4[1] - 14 * mm, settings.brand_name.upper())
            cv.setFont("Helvetica", 8)
            cv.setFillColor(MUTED)
            cv.drawRightString(A4[0] - doc.rightMargin, A4[1] - 14 * mm, self.company[:70])
            cv.setStrokeColor(LINE)
            cv.line(doc.leftMargin, A4[1] - 16 * mm, A4[0] - doc.rightMargin, A4[1] - 16 * mm)
        cv.setStrokeColor(LINE)
        cv.line(doc.leftMargin, 15 * mm, A4[0] - doc.rightMargin, 15 * mm)
        cv.setFont("Helvetica", 7.4)
        cv.setFillColor(MUTED)
        cv.drawString(doc.leftMargin, 11 * mm,
                      f"{settings.brand_name} · {settings.brand_owner} · {settings.contact_email}")
        cv.drawRightString(A4[0] - doc.rightMargin, 11 * mm, f"Page {page}")
        cv.restoreState()


# ---------------------------------------------------------------------------
#  Composants
# ---------------------------------------------------------------------------


def _radar_flowable(analysis: Analysis, width: float = 150 * mm, height: float = 105 * mm):
    """Radar dessiné directement sur le canevas PDF."""
    from reportlab.platypus import Flowable

    dims = analysis.score.dimensions
    labels = [d.short_name for d in dims]
    values = [d.percentage for d in dims]

    class Radar(Flowable):
        def wrap(self, *_args):
            return width, height

        def draw(self):
            cv = self.canv
            cx, cy, radius = width / 2, height / 2, min(width, height) * 0.36
            cv.saveState()
            for ring in range(1, 5):
                r = radius * ring / 4
                pts = axis_points(len(values), cx, cy, r)
                cv.setStrokeColor(LINE)
                cv.setLineWidth(0.5)
                path = cv.beginPath()
                path.moveTo(*pts[0])
                for p in pts[1:]:
                    path.lineTo(*p)
                path.close()
                cv.drawPath(path)
            cv.setStrokeColor(LINE)
            for x, y in axis_points(len(values), cx, cy, radius):
                cv.line(cx, cy, x, y)
            pts = radar_points(values, cx, cy, radius)
            path = cv.beginPath()
            path.moveTo(*pts[0])
            for p in pts[1:]:
                path.lineTo(*p)
            path.close()
            cv.setFillColor(colors.Color(0.31, 0.275, 0.898, alpha=0.22))
            cv.setStrokeColor(BRAND)
            cv.setLineWidth(1.6)
            cv.drawPath(path, stroke=1, fill=1)
            cv.setFillColor(BRAND)
            for x, y in pts:
                cv.circle(x, y, 1.9, stroke=0, fill=1)
            cv.setFont("Helvetica-Bold", 7.4)
            for index, label in enumerate(labels):
                lx, ly = axis_points(len(values), cx, cy, radius + 11 * mm)[index]
                cv.setFillColor(INK)
                cv.drawCentredString(lx, ly, label)
                cv.setFillColor(BRAND)
                cv.setFont("Helvetica", 7)
                cv.drawCentredString(lx, ly - 8, f"{values[index]:.0f} %")
                cv.setFont("Helvetica-Bold", 7.4)
            cv.restoreState()

    return Radar()


def _kpi_row(analysis: Analysis, st: dict) -> Table:
    score = analysis.score
    cells = [
        ("Score global", f"{score.total_score} / {score.max_score}"),
        ("Pourcentage", f"{score.percentage:.1f} %"),
        ("Niveau", score.level.name),
        ("Projection 12 mois", f"{analysis.projected_percentage:.0f} %"),
    ]
    data = [[Paragraph(label, st["small"]) for label, _ in cells],
            [Paragraph(value, st["kpi"]) for _, value in cells]]
    table = Table(data, colWidths=[43 * mm] * 4)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), SOFT),
        ("BOX", (0, 0), (-1, -1), 0.5, LINE),
        ("INNERGRID", (0, 0), (-1, -1), 0.5, colors.white),
        ("LEFTPADDING", (0, 0), (-1, -1), 8),
        ("RIGHTPADDING", (0, 0), (-1, -1), 8),
        ("TOPPADDING", (0, 0), (-1, -1), 7),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 9),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
    ]))
    return table


def _dimension_table(analysis: Analysis, st: dict) -> Table:
    headers = ["Dimension", "Poids", "Score", "Max", "%", "Niveau"]
    rows = [[Paragraph(h, st["cellh"] if index in (0, 5) else st["cellhr"])
             for index, h in enumerate(headers)]]
    for d in analysis.score.dimensions:
        rows.append([
            Paragraph(d.name, st["cell"]),
            Paragraph(str(d.weight), st["cellr"]),
            Paragraph(str(d.score), st["cellr"]),
            Paragraph(str(d.max_score), st["cellr"]),
            Paragraph(f"{d.percentage:.0f} %", st["cellr"]),
            Paragraph(d.level.name, st["cell"]),
        ])
    rows.append([
        Paragraph("Total", st["cellb"]),
        Paragraph("", st["cell"]),
        Paragraph(str(analysis.score.total_score), st["cellb"]),
        Paragraph(str(analysis.score.max_score), st["cellb"]),
        Paragraph(f"{analysis.score.percentage:.0f} %", st["cellb"]),
        Paragraph(analysis.score.level.name, st["cellb"]),
    ])
    table = Table(rows, colWidths=[62 * mm, 15 * mm, 20 * mm, 18 * mm, 18 * mm, 41 * mm],
                  repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), INK),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("BACKGROUND", (0, -1), (-1, -1), SOFT),
        ("GRID", (0, 0), (-1, -1), 0.4, LINE),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("LEFTPADDING", (0, 0), (-1, -1), 5),
        ("RIGHTPADDING", (0, 0), (-1, -1), 5),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]))
    for index, _ in enumerate(analysis.score.dimensions, start=1):
        if index % 2 == 0:
            table.setStyle(TableStyle([("BACKGROUND", (0, index), (-1, index),
                                       colors.HexColor("#FAFBFC"))]))
    return table


def _header_cells(st: dict, headers: List[str], right: tuple = ()) -> List[Paragraph]:
    """Cellules d'en-tête, en blanc sur le fond sombre du bandeau."""
    return [
        Paragraph(h, st["cellhr"] if index in right else st["cellh"])
        for index, h in enumerate(headers)
    ]


def _style_table(table: Table) -> Table:
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), INK),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("GRID", (0, 0), (-1, -1), 0.4, LINE),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("LEFTPADDING", (0, 0), (-1, -1), 5),
        ("RIGHTPADDING", (0, 0), (-1, -1), 5),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]))
    return table


# ---------------------------------------------------------------------------
#  Rapport principal
# ---------------------------------------------------------------------------


def build_report_pdf(
    assessment: Assessment,
    analysis: Analysis,
    *,
    plan_code: str = "premium",
    positioning=None,
) -> bytes:
    """Rapport complet. `plan_code` détermine les sections incluses.

    standard : synthèse, scores, analyse par dimension, recommandations
    premium  : ajoute la feuille de route 12 mois et la valeur en jeu
    """
    st = _styles()
    premium = plan_code == "premium"
    buffer = io.BytesIO()
    doc = _Doc(buffer, title=f"Diagnostic de maturité data — {assessment.company_name}",
               company=assessment.company_name)
    flow: List = []

    # ---- page de garde
    flow.append(Paragraph(settings.brand_name.upper(), st["h2"]))
    flow.append(Paragraph("Diagnostic de maturité data", st["title"]))
    flow.append(Paragraph(
        f"{assessment.company_name} — {assessment.sector} — {assessment.country}<br/>"
        f"Rapport {'Premium' if premium else 'Standard'} établi le "
        f"{datetime.now(timezone.utc).strftime('%d/%m/%Y')} "
        f"sur la base de la grille Limpida Consulting 2024 (45 critères, 7 dimensions, 768 points).",
        st["subtitle"]))
    flow.append(_kpi_row(analysis, st))
    flow.append(Spacer(1, 8 * mm))
    flow.append(_radar_flowable(analysis))
    flow.append(Spacer(1, 4 * mm))
    flow.append(Paragraph(
        f"Contact : {settings.brand_owner} · {settings.contact_email} · {settings.contact_phone}",
        st["small"]))
    flow.append(PageBreak())

    # ---- synthèse
    flow.append(Paragraph("1. Synthèse pour la direction", st["h1"]))
    flow.append(Paragraph(analysis.headline, st["h2"]))
    for para in analysis.executive_summary:
        flow.append(Paragraph(para, st["body"]))

    if positioning is not None and getattr(positioning, "comment", ""):
        flow.append(Paragraph("Positionnement sectoriel", st["h2"]))
        flow.append(Paragraph(positioning.comment, st["body"]))
        if positioning.sector and positioning.sector.is_publishable:
            flow.append(Paragraph(
                f"Moyenne du secteur {positioning.sector.label} : "
                f"{positioning.sector.average:.0f} % sur {positioning.sector.sample} organisations évaluées. "
                f"Médiane : {positioning.sector.median:.0f} %.", st["small"]))

    flow.append(Paragraph("2. Scores par dimension", st["h1"]))
    flow.append(_dimension_table(analysis, st))
    flow.append(Spacer(1, 3 * mm))
    flow.append(Paragraph(
        "Rappel de la méthode : score = réponse (0 à 3) x poids du critère x poids de la dimension. "
        "Les dimensions Gouvernance, Qualité, Sécurité et Infrastructure ont un poids de 3 : "
        "un même effort y produit un gain de score trois fois supérieur.", st["small"]))
    flow.append(PageBreak())

    # ---- analyse par dimension
    flow.append(Paragraph("3. Analyse détaillée par dimension", st["h1"]))
    for da in analysis.dimensions:
        block: List = [
            Paragraph(f"{da.result.name} — {da.result.percentage:.0f} % "
                      f"({da.result.score}/{da.result.max_score} points, poids {da.result.weight})",
                      st["h2"]),
            Paragraph(da.narrative, st["body"]),
        ]
        rows = [_header_cells(st, ["Critère", "Poids", "Niveau déclaré", "Score"], right=(1, 3))]
        for crit in da.result.criteria:
            rows.append([
                Paragraph(crit.name, st["cell"]),
                Paragraph(str(crit.weight), st["cellr"]),
                Paragraph(f"{crit.answer}/3 — {crit.level_label}", st["cell"]),
                Paragraph(f"{crit.score}/{crit.max_score}", st["cellr"]),
            ])
        block.append(_style_table(Table(rows, colWidths=[62 * mm, 15 * mm, 75 * mm, 22 * mm],
                                       repeatRows=1)))
        if premium and da.value_at_stake_xof:
            block.append(Spacer(1, 2 * mm))
            block.append(Paragraph(
                f"Valeur annuelle estimée en jeu sur cette dimension : "
                f"<b>{format_xof(da.value_at_stake_xof)}</b>.", st["small"]))
        block.append(Spacer(1, 5 * mm))
        flow.append(KeepTogether(block))

    # ---- recommandations
    flow.append(PageBreak())
    flow.append(Paragraph("4. Recommandations priorisées", st["h1"]))
    flow.append(Paragraph(
        "Les actions sont classées par rendement décroissant : priorité = poids de la dimension "
        "x poids du critère x écart au niveau maximum. Les premières lignes sont celles qui "
        "rapportent le plus de points de maturité par unité d'effort.", st["body"]))
    rows = [_header_cells(st, ["#", "Action", "Dimension", "Priorité", "Effort", "Levier", "Gain"],
                          right=(0, 6))]
    for reco in analysis.recommendations:
        rows.append([
            Paragraph(str(reco.priority_rank), st["cellr"]),
            Paragraph(reco.action, st["cell"]),
            Paragraph(reco.dimension_name.replace(" des données", ""), st["cell"]),
            Paragraph(reco.priority_label, st["cell"]),
            Paragraph(reco.effort.capitalize(), st["cell"]),
            Paragraph(reco.roi_driver_label, st["cell"]),
            Paragraph(f"+{reco.points_gain}", st["cellr"]),
        ])
    flow.append(_style_table(Table(
        rows, colWidths=[8 * mm, 58 * mm, 26 * mm, 18 * mm, 16 * mm, 28 * mm, 14 * mm],
        repeatRows=1)))

    # ---- détail opérationnel (premium)
    if premium:
        flow.append(PageBreak())
        flow.append(Paragraph("5. Mise en oeuvre des actions prioritaires", st["h1"]))
        for reco in analysis.recommendations[:12]:
            block = [
                Paragraph(f"{reco.priority_rank}. {reco.action}", st["h2"]),
                Paragraph(f"<b>Pourquoi :</b> {reco.why}", st["body"]),
                Paragraph(
                    "<b>Étapes :</b> " + " · ".join(f"({i+1}) {s}" for i, s in enumerate(reco.steps)),
                    st["body"]),
                Paragraph(
                    f"<b>Indicateur de suivi :</b> {reco.kpi}<br/>"
                    f"<b>Niveau actuel :</b> {reco.current_label or 'non renseigné'}<br/>"
                    f"<b>Niveau visé :</b> {reco.target_label}<br/>"
                    f"<b>Trimestre :</b> T{reco.quarter} · <b>Effort :</b> {reco.effort} · "
                    f"<b>Levier :</b> {reco.roi_driver_label}", st["small"]),
                Spacer(1, 3 * mm),
            ]
            flow.append(KeepTogether(block))

        flow.append(PageBreak())
        flow.append(Paragraph("6. Feuille de route 12 mois", st["h1"]))
        for quarter in analysis.roadmap:
            if not quarter.recommendations:
                continue
            rows = [_header_cells(st, ["Action", "Dimension", "Effort", "Gain de points"], right=(3,))]
            for reco in quarter.recommendations:
                rows.append([
                    Paragraph(reco.action, st["cell"]),
                    Paragraph(reco.dimension_name.replace(" des données", ""), st["cell"]),
                    Paragraph(reco.effort.capitalize(), st["cell"]),
                    Paragraph(f"+{reco.points_gain}", st["cellr"]),
                ])
            block = [
                Paragraph(f"{quarter.label} — {quarter.theme} "
                          f"({len(quarter.recommendations)} actions, +{quarter.points_gain} points)",
                          st["h2"]),
                _style_table(Table(rows, colWidths=[85 * mm, 40 * mm, 22 * mm, 21 * mm],
                                   repeatRows=1)),
                Spacer(1, 4 * mm),
            ]
            flow.append(KeepTogether(block))

        flow.append(PageBreak())
        flow.append(Paragraph("7. Valeur en jeu et hypothèses", st["h1"]))
        rows = [_header_cells(st, ["Dimension", "Écart de maturité", "Valeur annuelle estimée"],
                              right=(1, 2))]
        for da in sorted(analysis.dimensions, key=lambda d: -d.value_at_stake_xof):
            rows.append([
                Paragraph(da.result.name, st["cell"]),
                Paragraph(f"{100 - da.result.percentage:.0f} points", st["cellr"]),
                Paragraph(format_xof(da.value_at_stake_xof), st["cellr"]),
            ])
        rows.append([
            Paragraph("Total", st["cellb"]),
            Paragraph("", st["cell"]),
            Paragraph(format_xof(analysis.total_value_at_stake_xof), st["cellb"]),
        ])
        flow.append(_style_table(Table(rows, colWidths=[80 * mm, 45 * mm, 43 * mm], repeatRows=1)))
        flow.append(Spacer(1, 4 * mm))
        flow.append(Paragraph("Hypothèses de calcul", st["h2"]))
        for item in analysis.assumptions:
            flow.append(Paragraph(f"• {item}", st["small"]))

        if analysis.quick_wins:
            flow.append(Spacer(1, 5 * mm))
            flow.append(Paragraph("Actions à effort faible à lancer immédiatement", st["h2"]))
            for reco in analysis.quick_wins:
                flow.append(Paragraph(f"• {reco.action} — {reco.dimension_name}", st["small"]))

    flow.append(Spacer(1, 8 * mm))
    flow.append(Paragraph(
        f"Rapport établi par {settings.brand_owner} — {settings.brand_name}. "
        f"Grille de référence : Limpida Consulting 2024. "
        f"Ce document est destiné à l'usage interne de {assessment.company_name}.", st["small"]))

    doc.build(flow)
    return buffer.getvalue()


# ---------------------------------------------------------------------------
#  Export Excel de la grille complétée
# ---------------------------------------------------------------------------


def build_grid_xlsx(
    assessment: Assessment, analysis: Analysis, *, include_action_plan: bool = True
) -> bytes:
    """Grille complétée, réutilisable par le client dans ses propres travaux.

    `include_action_plan=False` produit la version gratuite : synthèse et grille
    détaillée, sans le plan d'action qui relève de l'offre payante.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    thin = Side(style="thin", color="DDE3EA")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    head_fill = PatternFill("solid", fgColor="0F172A")
    head_font = Font(bold=True, color="FFFFFF", size=10)
    bold = Font(bold=True, size=10)

    # -- feuille 1 : synthèse
    ws = wb.active
    ws.title = "Synthèse"
    ws["A1"] = f"{settings.brand_name} — Diagnostic de maturité data"
    ws["A1"].font = Font(bold=True, size=14)
    meta = [
        ("Organisation", assessment.company_name),
        ("Secteur", assessment.sector),
        ("Pays", assessment.country),
        ("Effectif", assessment.company_size),
        ("Contact", f"{assessment.contact_name} — {assessment.contact_email}"),
        ("Date d'évaluation", (assessment.completed_at or assessment.created_at).strftime("%d/%m/%Y")),
        ("Score global", f"{analysis.score.total_score} / {analysis.score.max_score}"),
        ("Pourcentage", f"{analysis.score.percentage:.1f} %"),
        ("Niveau de maturité", analysis.score.level.name),
        ("Projection à 12 mois", f"{analysis.projected_percentage:.0f} %"),
    ]
    for index, (label, value) in enumerate(meta, start=3):
        ws.cell(row=index, column=1, value=label).font = bold
        ws.cell(row=index, column=2, value=value)
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 58

    row = len(meta) + 5
    ws.cell(row=row, column=1, value="Dimension").font = head_font
    for col, header in enumerate(["Dimension", "Poids", "Score", "Maximum", "%", "Niveau"], start=1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font, cell.fill, cell.border = head_font, head_fill, border
    for d in analysis.score.dimensions:
        row += 1
        for col, value in enumerate(
            [d.name, d.weight, d.score, d.max_score, d.percentage / 100, d.level.name], start=1
        ):
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = border
            if col == 5:
                cell.number_format = "0.0%"
    row += 1
    for col, value in enumerate(
        ["Total", "", analysis.score.total_score, analysis.score.max_score,
         analysis.score.percentage / 100, analysis.score.level.name], start=1
    ):
        cell = ws.cell(row=row, column=col, value=value)
        cell.font, cell.border = bold, border
        if col == 5:
            cell.number_format = "0.0%"
    for col, width in enumerate([34, 8, 10, 10, 10, 16], start=1):
        ws.column_dimensions[get_column_letter(col)].width = width

    # -- feuille 2 : grille détaillée
    ws2 = wb.create_sheet("Grille détaillée")
    headers = ["Dimension", "Poids dimension", "Critère", "Poids critère",
               "Réponse (0-3)", "Niveau déclaré", "Score", "Score maximum"]
    for col, header in enumerate(headers, start=1):
        cell = ws2.cell(row=1, column=col, value=header)
        cell.font, cell.fill, cell.border = head_font, head_fill, border
        cell.alignment = Alignment(wrap_text=True, vertical="center")
    row = 2
    for dim in analysis.score.dimensions:
        for crit in dim.criteria:
            values = [dim.name, dim.weight, crit.name, crit.weight, crit.answer,
                      crit.level_label, crit.score, crit.max_score]
            for col, value in enumerate(values, start=1):
                cell = ws2.cell(row=row, column=col, value=value)
                cell.border = border
                cell.alignment = Alignment(wrap_text=(col in {1, 3, 6}), vertical="top")
            row += 1
    for col, width in enumerate([26, 12, 40, 10, 12, 52, 9, 13], start=1):
        ws2.column_dimensions[get_column_letter(col)].width = width
    ws2.freeze_panes = "A2"

    # -- feuille 3 : plan d'action (offre payante uniquement)
    if not include_action_plan:
        buffer = io.BytesIO()
        wb.save(buffer)
        return buffer.getvalue()

    ws3 = wb.create_sheet("Plan d'action")
    headers = ["Rang", "Trimestre", "Dimension", "Action", "Priorité", "Effort",
               "Levier", "Gain de points", "Indicateur de suivi"]
    for col, header in enumerate(headers, start=1):
        cell = ws3.cell(row=1, column=col, value=header)
        cell.font, cell.fill, cell.border = head_font, head_fill, border
        cell.alignment = Alignment(wrap_text=True, vertical="center")
    for row, reco in enumerate(analysis.recommendations, start=2):
        values = [reco.priority_rank, f"T{reco.quarter}", reco.dimension_name, reco.action,
                  reco.priority_label, reco.effort, reco.roi_driver_label,
                  reco.points_gain, reco.kpi]
        for col, value in enumerate(values, start=1):
            cell = ws3.cell(row=row, column=col, value=value)
            cell.border = border
            cell.alignment = Alignment(wrap_text=(col in {3, 4, 9}), vertical="top")
    for col, width in enumerate([7, 11, 26, 46, 12, 10, 22, 13, 42], start=1):
        ws3.column_dimensions[get_column_letter(col)].width = width
    ws3.freeze_panes = "A2"

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def report_filename(assessment: Assessment, extension: str, plan_code: str = "premium") -> str:
    safe = "".join(ch if ch.isalnum() or ch in "-_" else "-" for ch in assessment.company_name)[:40]
    stamp = (assessment.completed_at or assessment.created_at).strftime("%Y%m%d")
    return f"DataMaturity-{plan_code}-{safe or 'organisation'}-{stamp}.{extension}"
